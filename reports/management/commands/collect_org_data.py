"""
Собрать максимум реквизитов по списку ИНН из доступных источников.

Источники:
  - БД Certificate / Organization
  - CSV auth-скана (статус, сроки ЭЦП)
  - Subject сертификата (ФИО, КПП)
  - СБИС.СписокНашихОрганизаций (после auth по серту)
  - star-pro.ru по ИНН (КПП, ОГРН, краткое имя) — опционально

Пример:
  docker compose exec web python manage.py collect_org_data --from-file valid_inns_final.txt --limit 5
  docker compose exec web python manage.py collect_org_data --from-file valid_inns_final.txt --sbis --workers 4 --quiet
  docker compose exec web python manage.py collect_org_data \\
    --from-file valid_inns_final.txt \\
    --auth-csv /app/media/sbis_auth_scan/sbis_auth_report_YYYYMMDD_HHMMSS.csv \\
    --sbis --workers 4 --quiet

Мониторинг: /app/media/org_export/collect_org_LIVE.log и collect_org_LIVE.status.json
"""
from __future__ import annotations

import csv
import json
import logging
import os
import sys
import tempfile
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import close_old_connections
from django.utils import timezone as dj_timezone

from reports.management.commands.test_sbis_auth_all import (
    _load_inns_from_file,
    classify_sbis_error,
)
from reports.models import Certificate, Organization
from reports.services.kpp_external import fetch_kpp_star_pro
from reports.services.sbis import export_cert_der, parse_kpp_from_cert_file
from reports.services.sbis.organizations import (
    kpp_to_tax_office_code,
    pick_best_sbis_org,
    sbis_list_our_organizations,
)
from reports.services.sbis_mail import SbisAuthError, SbisSessionService

# Колонки для передачи в 1С / внешние файлы
EXPORT_COLUMNS = [
    "ИНН",
    "КПП",
    "ОГРН",
    "Наименование",
    "НаименованиеПолное",
    "НаименованиеСокращенное",
    "КодНалоговогоОргана",
    "ЮрлицоИлиИП",
    "ФамилияИП",
    "ИмяИП",
    "ОтчествоИП",
    "ЭЦП_действует_с",
    "ЭЦП_действует_по",
    "ЭЦПОтозвана",
]


def _empty_export_row(inn: str = "") -> dict:
    row = {col: "" for col in EXPORT_COLUMNS}
    if inn:
        row["ИНН"] = inn
    row["ЭЦПОтозвана"] = "нет"
    return row


def _fmt_dt(dt) -> str:
    if not dt:
        return ""
    if dj_timezone.is_naive(dt):
        dt = dj_timezone.make_aware(dt, timezone.utc)
    return dt.astimezone(timezone.utc).strftime("%Y-%m-%d")


def _load_auth_csv(path: Path) -> dict[str, dict]:
    if not path.is_file():
        return {}
    out: dict[str, dict] = {}
    with path.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            inn = (row.get("inn") or "").strip()
            if inn:
                out[inn] = row
    return out


def _first_non_empty(*values: str) -> str:
    for v in values:
        s = (v or "").strip()
        if s:
            return s
    return ""


def _parse_csv_date(val: str) -> str:
    s = (val or "").strip()
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    return s[:10] if s else ""


def _collect_one(
    inn: str,
    *,
    auth_by_inn: dict[str, dict],
    call_sbis: bool,
    call_egrul: bool,
    parse_cert: bool,
    proxy_want: int,
    proxy_budget: int,
) -> tuple[dict, dict]:
    """Возвращает (строка для экспорта, служебная meta для логов)."""
    close_old_connections()
    row = _empty_export_row(inn)
    meta = {
        "auth_status": "неизвестно",
        "error_category": "",
        "sbis_error": "",
        "egrul_error": "",
    }

    auth_row = auth_by_inn.get(inn) or {}
    if auth_row:
        status = (auth_row.get("status") or "").strip()
        cat = (auth_row.get("error_category") or "").strip()
        meta["auth_status"] = "ok" if status == "ok" else "fail"
        meta["error_category"] = cat
        if cat == "revoked_or_untrusted":
            row["ЭЦПОтозвана"] = "да"
        row["ЭЦП_действует_с"] = _parse_csv_date(auth_row.get("not_before") or "")
        row["ЭЦП_действует_по"] = _parse_csv_date(auth_row.get("not_after") or "")

    cert = (
        Certificate.objects.filter(inn=inn, has_private_key=True, is_active=True)
        .order_by("-not_after", "-id")
        .first()
    )
    org = Organization.objects.filter(inn=inn).first()

    if cert:
        if not row["ЭЦП_действует_с"] and cert.not_before:
            row["ЭЦП_действует_с"] = _fmt_dt(cert.not_before)
        if not row["ЭЦП_действует_по"] and cert.not_after:
            row["ЭЦП_действует_по"] = _fmt_dt(cert.not_after)
        if cert.kpp and not row["КПП"]:
            row["КПП"] = str(cert.kpp).strip()

        if parse_cert and cert.csptest_name:
            fd, cert_path = tempfile.mkstemp(prefix=f"org_col_{inn}_", suffix=".cer")
            os.close(fd)
            try:
                export_cert_der(cert.csptest_name, cert_path)
                cert_kpp = parse_kpp_from_cert_file(cert_path) or ""
                if cert_kpp and not row["КПП"]:
                    row["КПП"] = cert_kpp
            except Exception:
                pass
            finally:
                try:
                    os.remove(cert_path)
                except OSError:
                    pass

    if org:
        if org.kpp and not row["КПП"]:
            row["КПП"] = org.kpp.strip()
        if org.name and not row["Наименование"]:
            row["Наименование"] = org.name.strip()

    if call_egrul:
        try:
            egr = fetch_kpp_star_pro(inn)
            if egr.get("ok"):
                if egr.get("kpp") and not row["КПП"]:
                    row["КПП"] = egr["kpp"]
                if egr.get("ogrn"):
                    row["ОГРН"] = egr["ogrn"]
                if egr.get("name_short"):
                    row["НаименованиеСокращенное"] = egr["name_short"]
                    if not row["Наименование"]:
                        row["Наименование"] = egr["name_short"]
            else:
                meta["egrul_error"] = (egr.get("error") or "unknown")[:200]
        except Exception as e:
            meta["egrul_error"] = str(e)[:200]

    pack: dict = {"success": False, "organizations": []}
    if call_sbis:
        if not cert:
            meta["sbis_error"] = "нет активного сертификата в БД"
        else:
            try:
                session_id = SbisSessionService(
                    certificate=cert,
                    proxy_want=proxy_want,
                    proxy_warmup_budget_sec=proxy_budget,
                ).authenticate()
                meta["auth_status"] = "ok"
                pack = sbis_list_our_organizations(
                    inn,
                    session_id,
                    filter_inn=inn,
                    filter_kpp=row["КПП"] or "",
                )
            except SbisAuthError as e:
                meta["sbis_error"] = str(e)[:300]
                meta["auth_status"] = "fail"
                meta["error_category"] = classify_sbis_error(str(e))
            except Exception as e:
                meta["sbis_error"] = str(e)[:300]

    if pack.get("success"):
        best = pick_best_sbis_org(pack.get("organizations") or [], inn)
        if best:
            if best.get("kpp"):
                row["КПП"] = _first_non_empty(best["kpp"], row["КПП"])
            if best.get("name"):
                row["Наименование"] = _first_non_empty(best["name"], row["Наименование"])
                row["НаименованиеПолное"] = best["name"]
            et = best.get("entity_type") or ""
            row["ЮрлицоИлиИП"] = "ИП" if et == "IP" else ("ЮЛ" if et == "UL" else "")
            if et == "IP":
                row["ФамилияИП"] = best.get("surname") or ""
                row["ИмяИП"] = best.get("firstname") or ""
                row["ОтчествоИП"] = best.get("patronymic") or ""
    elif pack.get("error") and not meta["sbis_error"]:
        err = pack["error"]
        meta["sbis_error"] = str(err.get("message") if isinstance(err, dict) else err)[:300]

    if not row["НаименованиеПолное"] and row["Наименование"]:
        row["НаименованиеПолное"] = row["Наименование"]
    if not row["НаименованиеСокращенное"] and row["Наименование"]:
        row["НаименованиеСокращенное"] = row["Наименование"]

    if row["КПП"] and not row["КодНалоговогоОргана"]:
        row["КодНалоговогоОргана"] = kpp_to_tax_office_code(row["КПП"])

    if len(inn) == 12:
        row["ЮрлицоИлиИП"] = row["ЮрлицоИлиИП"] or "ИП"
    elif len(inn) == 10:
        row["ЮрлицоИлиИП"] = row["ЮрлицоИлиИП"] or "ЮЛ"

    return row, meta


class Command(BaseCommand):
    help = "Собрать максимум реквизитов организаций для выгрузки в 1С"

    def add_arguments(self, parser):
        parser.add_argument(
            "--from-file",
            default="",
            help="Список ИНН (по одному в строке), напр. valid_inns_final.txt",
        )
        parser.add_argument(
            "--auth-csv",
            default="",
            help="CSV из test_sbis_auth_all (статусы и сроки ЭЦП)",
        )
        parser.add_argument(
            "--output-dir",
            default="/app/media/org_export",
            help="Каталог для JSON/CSV/XLSX",
        )
        parser.add_argument("--limit", type=int, default=0)
        parser.add_argument("--offset", type=int, default=0)
        parser.add_argument(
            "--sbis",
            action="store_true",
            help="СБИС.СписокНашихОрганизаций (auth + HTTP на каждый ИНН)",
        )
        parser.add_argument(
            "--no-egrul",
            action="store_true",
            help="Не запрашивать star-pro.ru (КПП/ОГРН/имя). По умолчанию egrul включён",
        )
        parser.add_argument(
            "--egrul",
            action="store_true",
            help="(устар.) то же что по умолчанию; оставлено для совместимости",
        )
        parser.add_argument(
            "--parse-cert",
            action="store_true",
            help="Экспорт .cer через certmgr на каждый ИНН (медленно, для КПП из Subject)",
        )
        parser.add_argument("--egrul-delay", type=float, default=0.35)
        parser.add_argument("--workers", type=int, default=1)
        parser.add_argument("--proxy-want", type=int, default=2)
        parser.add_argument("--proxy-budget", type=int, default=8)
        parser.add_argument(
            "--per-inn-json",
            action="store_true",
            help="Положить отдельный {ИНН}.json в подкаталог companies/",
        )
        parser.add_argument("--quiet", action="store_true")
        parser.add_argument(
            "--progress-every",
            type=int,
            default=0,
            help="Строка в лог каждые N ИНН (0=авто: 1 при --quiet, иначе 10)",
        )

    def _write_csv(self, path: Path, rows: list[dict]) -> None:
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=EXPORT_COLUMNS, extrasaction="ignore")
            w.writeheader()
            w.writerows(rows)

    def _write_status(
        self,
        path: Path,
        *,
        state: str,
        total: int,
        rows: list[dict],
        meta_rows: list[dict],
        started_at: str,
        last_inn: str = "",
        last_row: dict | None = None,
        extra: dict | None = None,
    ) -> None:
        done = len(rows)
        with_kpp = sum(1 for r in rows if r.get("КПП"))
        with_name = sum(1 for r in rows if r.get("Наименование"))
        sbis_err = sum(1 for m in meta_rows if m.get("sbis_error"))
        egrul_err = sum(1 for m in meta_rows if m.get("egrul_error"))
        payload = {
            "state": state,
            "done": done,
            "total": total,
            "percent": round(100.0 * done / total, 1) if total else 0,
            "with_kpp": with_kpp,
            "with_name": with_name,
            "sbis_errors": sbis_err,
            "egrul_errors": egrul_err,
            "started_at": started_at,
            "updated_at": dj_timezone.now().isoformat(),
            "last_inn": last_inn,
            "last": last_row,
        }
        if extra:
            payload.update(extra)
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def handle(self, *args, **options):
        if options["quiet"]:
            logging.getLogger("reports.services.sbis").setLevel(logging.WARNING)

        inns: list[str] = []
        if options.get("from_file"):
            inns = _load_inns_from_file(Path(options["from_file"]))
        if not inns:
            inns = sorted(
                Certificate.objects.filter(has_private_key=True, is_active=True)
                .exclude(inn="")
                .values_list("inn", flat=True)
                .distinct()
            )

        if options["offset"]:
            inns = inns[options["offset"] :]
        if options["limit"]:
            inns = inns[: options["limit"]]

        auth_csv = Path(options["auth_csv"]) if options.get("auth_csv") else None
        if auth_csv is None:
            for candidate in (
                Path("sbis_auth_report_20260618_133531.csv"),
                Path("/app/media/sbis_auth_scan/sbis_auth_report_20260618_133531.csv"),
            ):
                if candidate.is_file():
                    auth_csv = candidate
                    break
        auth_by_inn = _load_auth_csv(auth_csv) if auth_csv else {}

        out_dir = Path(options["output_dir"])
        out_dir.mkdir(parents=True, exist_ok=True)
        ts = dj_timezone.now().strftime("%Y%m%d_%H%M%S")
        json_path = out_dir / f"organizations_{ts}.json"
        csv_path = out_dir / f"organizations_{ts}.csv"
        progress_csv_path = out_dir / f"organizations_{ts}_progress.csv"
        live_status_path = out_dir / "collect_org_LIVE.status.json"
        live_log_path = out_dir / "collect_org_LIVE.log"
        live_csv_path = out_dir / "collect_org_LIVE_progress.csv"
        xlsx_path = out_dir / f"organizations_{ts}.xlsx"
        companies_dir = out_dir / f"companies_{ts}"

        workers = max(1, min(8, int(options.get("workers") or 1)))
        call_sbis = bool(options["sbis"])
        call_egrul = not bool(options["no_egrul"])
        parse_cert = bool(options["parse_cert"])
        egrul_delay = max(0.0, float(options["egrul_delay"]))
        pe = int(options.get("progress_every") or 0)
        progress_every = pe if pe > 0 else (1 if options["quiet"] else 10)
        started_at = dj_timezone.now().isoformat()

        self.stdout.write(f"ИНН к сбору: {len(inns)}")
        self.stdout.write(f"СБИС СписокНашихОрганизаций: {'да' if call_sbis else 'нет'}")
        self.stdout.write(f"star-pro (КПП/ОГРН/имя): {'да' if call_egrul else 'нет'}")
        self.stdout.write(f"Парсинг сертификата (certmgr): {'да' if parse_cert else 'нет'}")
        self.stdout.write(f"Потоков: {workers}")
        self.stdout.write(f"Лог прогресса:  {live_log_path}")
        self.stdout.write(f"Статус live:    {live_status_path}")
        self.stdout.write(f"CSV live:       {live_csv_path}")
        self.stdout.write(f"Строка в лог каждые {progress_every} ИНН")
        sys.stdout.flush()
        if auth_csv:
            self.stdout.write(f"Auth CSV: {auth_csv}")
        if not call_sbis and not call_egrul:
            self.stdout.write(
                self.style.WARNING(
                    "Без --sbis и --no-egrul в файле будут в основном ИНН и сроки ЭЦП из CSV"
                )
            )

        rows: list[dict] = []
        meta_rows: list[dict] = []
        lock = threading.Lock()
        done = [0]
        total = len(inns)

        self._write_status(
            live_status_path,
            state="running",
            total=total,
            rows=rows,
            meta_rows=meta_rows,
            started_at=started_at,
            extra={
                "sbis": call_sbis,
                "egrul": call_egrul,
                "workers": workers,
                "pid": os.getpid(),
            },
        )
        live_log_path.write_text(
            f"START {started_at} total={total} sbis={call_sbis} egrul={call_egrul} workers={workers}\n",
            encoding="utf-8",
        )

        def _run(inn: str) -> tuple[dict, dict]:
            if call_egrul and egrul_delay > 0:
                time.sleep(egrul_delay)
            return _collect_one(
                inn,
                auth_by_inn=auth_by_inn,
                call_sbis=call_sbis,
                call_egrul=call_egrul,
                parse_cert=parse_cert,
                proxy_want=max(1, int(options["proxy_want"])),
                proxy_budget=max(4, int(options["proxy_budget"])),
            )

        def _store(r: dict, meta: dict) -> None:
            with lock:
                rows.append(r)
                meta_rows.append(meta)
                done[0] += 1
                n = done[0]
                line = (
                    f"[{n}/{total}] {r['ИНН']} — "
                    f"auth={meta['auth_status']}, "
                    f"ЭЦП до {r['ЭЦП_действует_по'] or '—'}, "
                    f"КПП={r['КПП'] or '—'}, "
                    f"имя={(r['Наименование'] or '—')[:35]}"
                )
                if meta.get("sbis_error"):
                    line += f" | sbis_err={(meta['sbis_error'] or '')[:60]}"
                if meta.get("egrul_error"):
                    line += f" | egrul_err={(meta['egrul_error'] or '')[:40]}"

                with live_log_path.open("a", encoding="utf-8") as lf:
                    lf.write(line + "\n")

                show = (not options["quiet"]) or (n % progress_every == 0) or (n == total)
                if show:
                    self.stdout.write(line)
                    sys.stdout.flush()

                snapshot = sorted(rows, key=lambda x: x.get("ИНН", ""))
                self._write_csv(live_csv_path, snapshot)
                if n % progress_every == 0 or n == total:
                    self._write_csv(progress_csv_path, snapshot)

                self._write_status(
                    live_status_path,
                    state="running",
                    total=total,
                    rows=rows,
                    meta_rows=meta_rows,
                    started_at=started_at,
                    last_inn=r["ИНН"],
                    last_row=r,
                )

                if options["per_inn_json"]:
                    companies_dir.mkdir(parents=True, exist_ok=True)
                    (companies_dir / f"{r['ИНН']}.json").write_text(
                        json.dumps(r, ensure_ascii=False, indent=2),
                        encoding="utf-8",
                    )

        if workers <= 1:
            for inn in inns:
                row, meta = _run(inn)
                _store(row, meta)
        else:
            with ThreadPoolExecutor(max_workers=workers) as pool:
                futs = {pool.submit(_run, inn): inn for inn in inns}
                for fut in as_completed(futs):
                    try:
                        row, meta = fut.result()
                        _store(row, meta)
                    except Exception as e:
                        inn = futs[fut]
                        _store(
                            _empty_export_row(inn),
                            {"auth_status": "fail", "error_category": "", "sbis_error": str(e), "egrul_error": ""},
                        )

        rows.sort(key=lambda r: r.get("ИНН", ""))

        json_path.write_text(
            json.dumps(rows, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=EXPORT_COLUMNS, extrasaction="ignore")
            w.writeheader()
            w.writerows(rows)

        try:
            progress_csv_path.unlink(missing_ok=True)
        except OSError:
            pass

        if options["per_inn_json"]:
            companies_dir.mkdir(parents=True, exist_ok=True)
            for r in rows:
                (companies_dir / f"{r['ИНН']}.json").write_text(
                    json.dumps(r, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Организации"
            ws.append(EXPORT_COLUMNS)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1565C0")
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            for r in rows:
                ws.append([r.get(c, "") for c in EXPORT_COLUMNS])
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for i, h in enumerate(EXPORT_COLUMNS, 1):
                ws.column_dimensions[get_column_letter(i)].width = min(max(len(h) + 2, 12), 36)
            wb.save(xlsx_path)
            xlsx_msg = str(xlsx_path)
        except ImportError:
            xlsx_msg = "(openpyxl не установлен — только JSON/CSV)"

        self._write_status(
            live_status_path,
            state="done",
            total=total,
            rows=rows,
            meta_rows=meta_rows,
            started_at=started_at,
            extra={"json": str(json_path), "csv": str(csv_path)},
        )
        with live_log_path.open("a", encoding="utf-8") as lf:
            lf.write(f"DONE {dj_timezone.now().isoformat()} records={len(rows)}\n")

        filled_kpp = sum(1 for r in rows if r.get("КПП"))
        filled_name = sum(1 for r in rows if r.get("Наименование"))
        filled_ogrn = sum(1 for r in rows if r.get("ОГРН"))

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS(f"Записей: {len(rows)}"))
        self.stdout.write(f"  с КПП:         {filled_kpp}")
        self.stdout.write(f"  с наименованием: {filled_name}")
        self.stdout.write(f"  с ОГРН:        {filled_ogrn}")
        self.stdout.write(f"JSON:  {json_path}")
        self.stdout.write(f"CSV:   {csv_path}")
        self.stdout.write(f"Excel: {xlsx_msg}")
        if options["per_inn_json"]:
            self.stdout.write(f"JSON/ИНН: {companies_dir}")
